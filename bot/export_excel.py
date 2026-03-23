"""Excel report generator for betting-sim.

Generates a dated .xlsx file with two sheets:
- Bets: all bets from the database
- Daily Summary: one row per day with bankroll stats
"""

import os
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ── colour constants ──────────────────────────────────────────────────────────
HEADER_BG = "9DC3E6"
WON_BG = "C6EFCE"
LOST_BG = "FFC7CE"
VOID_BG = "FFEB9C"

RESULT_FILL = {
    "won": PatternFill("solid", fgColor=WON_BG),
    "lost": PatternFill("solid", fgColor=LOST_BG),
    "void": PatternFill("solid", fgColor=VOID_BG),
}

HEADER_FILL = PatternFill("solid", fgColor=HEADER_BG)
HEADER_FONT = Font(bold=True)


# ── helpers ───────────────────────────────────────────────────────────────────

def _autofit(ws) -> None:
    """Estimate and set column widths based on cell contents."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                cell_len = len(str(cell.value)) if cell.value is not None else 0
                if cell_len > max_len:
                    max_len = cell_len
            except Exception:
                pass
        # Add a small padding; cap at 50 chars
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


def _apply_header(ws, headers: list[str]) -> None:
    """Write bold, light-blue header row."""
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


# ── sheet builders ────────────────────────────────────────────────────────────

def _build_bets_sheet(ws, bets: list[dict]) -> None:
    headers = [
        "Date", "Sport", "Event", "Market", "Selection",
        "Odds", "Implied Prob", "Stake ($)", "Result", "P&L ($)",
    ]
    _apply_header(ws, headers)
    ws.freeze_panes = "A2"

    for row_idx, bet in enumerate(bets, start=2):
        # Date
        raw_date = bet.get("date", "")
        date_val = raw_date[:10] if raw_date else ""  # keep YYYY-MM-DD
        ws.cell(row=row_idx, column=1, value=date_val)

        ws.cell(row=row_idx, column=2, value=bet.get("sport", ""))
        ws.cell(row=row_idx, column=3, value=bet.get("event_name", ""))
        ws.cell(row=row_idx, column=4, value=bet.get("market", ""))
        ws.cell(row=row_idx, column=5, value=bet.get("selection", ""))

        # Odds
        odds = bet.get("decimal_odds")
        ws.cell(row=row_idx, column=6, value=float(odds) if odds is not None else None)

        # Implied Prob — store as decimal, format as percentage
        implied = bet.get("implied_prob")
        prob_cell = ws.cell(row=row_idx, column=7,
                            value=float(implied) if implied is not None else None)
        if implied is not None:
            prob_cell.number_format = "0.0%"

        # Stake
        stake = bet.get("stake")
        stake_cell = ws.cell(row=row_idx, column=8,
                             value=float(stake) if stake is not None else None)
        if stake is not None:
            stake_cell.number_format = "#,##0.00"

        # Result
        result = (bet.get("result") or "pending").lower()
        result_cell = ws.cell(row=row_idx, column=9, value=result)
        fill = RESULT_FILL.get(result)
        if fill:
            result_cell.fill = fill

        # P&L
        pl = bet.get("profit_loss")
        pl_cell = ws.cell(row=row_idx, column=10,
                          value=float(pl) if pl is not None else None)
        if pl is not None:
            pl_cell.number_format = "#,##0.00"

    _autofit(ws)


def _build_daily_summary_sheet(ws, history: list[dict]) -> None:
    headers = ["Date", "Bankroll ($)", "Day P&L ($)", "Win Rate", "Bets"]
    _apply_header(ws, headers)
    ws.freeze_panes = "A2"

    total_pl = 0.0
    total_won = 0
    total_bets = 0

    for row_idx, record in enumerate(history, start=2):
        raw_date = record.get("date", "")
        date_val = raw_date[:10] if raw_date else ""
        ws.cell(row=row_idx, column=1, value=date_val)

        bankroll = record.get("closing_balance")
        ws.cell(row=row_idx, column=2,
                value=float(bankroll) if bankroll is not None else None).number_format = "#,##0.00"

        daily_pl = record.get("daily_pl")
        ws.cell(row=row_idx, column=3,
                value=float(daily_pl) if daily_pl is not None else None).number_format = "#,##0.00"

        # Compute win rate from num_won / num_bets if available; else 0
        num_won = record.get("num_won", 0) or 0
        num_bets = record.get("num_bets", 0) or 0
        win_rate = (num_won / num_bets) if num_bets > 0 else 0.0
        wr_cell = ws.cell(row=row_idx, column=4, value=win_rate)
        wr_cell.number_format = "0.0%"

        ws.cell(row=row_idx, column=5, value=num_bets)

        # Accumulators for totals row
        if daily_pl is not None:
            total_pl += float(daily_pl)
        total_won += num_won
        total_bets += num_bets

    # Totals row
    totals_row = len(history) + 2
    total_cell = ws.cell(row=totals_row, column=1, value="TOTAL")
    total_cell.font = HEADER_FONT

    ws.cell(row=totals_row, column=2, value=None)  # bankroll total doesn't aggregate simply

    total_pl_cell = ws.cell(row=totals_row, column=3, value=total_pl)
    total_pl_cell.number_format = "#,##0.00"
    total_pl_cell.font = HEADER_FONT

    overall_wr = (total_won / total_bets) if total_bets > 0 else 0.0
    wr_total_cell = ws.cell(row=totals_row, column=4, value=overall_wr)
    wr_total_cell.number_format = "0.0%"
    wr_total_cell.font = HEADER_FONT

    ws.cell(row=totals_row, column=5, value=total_bets).font = HEADER_FONT

    _autofit(ws)


# ── public API ────────────────────────────────────────────────────────────────

def generate_excel(output_dir: str = ".") -> str:
    """Generate Excel report. Returns absolute file path."""
    from bot.db import get_all_bets, get_bankroll_history

    bets = get_all_bets()
    history = get_bankroll_history()

    wb = Workbook()

    # Sheet 1 — Bets
    ws_bets = wb.active
    ws_bets.title = "Bets"
    _build_bets_sheet(ws_bets, bets)

    # Sheet 2 — Daily Summary
    ws_summary = wb.create_sheet("Daily Summary")
    _build_daily_summary_sheet(ws_summary, history)

    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, f"betting_sim_{date.today().isoformat()}.xlsx")
    wb.save(filepath)
    return os.path.abspath(filepath)


if __name__ == "__main__":
    path = generate_excel()
    print(f"Excel saved to: {path}")
